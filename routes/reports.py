from datetime import datetime
from flask import Blueprint, send_file, request, redirect, url_for
from flask_login import login_required, current_user
from sqlalchemy import func
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app import db
from models import PaymentRequest, Branch

reports_bp = Blueprint("reports", __name__)


@reports_bp.route("/reports")
@login_required
def reports():
    # Reports are now embedded in the dashboard — redirect there
    return redirect(url_for("requests.dashboard", **request.args))


@reports_bp.route("/export")
@login_required
def export():
    month      = request.args.get("month", "")
    branch_id  = request.args.get("branch_id", type=int)
    status_f   = request.args.get("status", "approved")
    inc_upload = request.args.get("include_uploaded", "0") == "1"

    q = PaymentRequest.query.filter_by(status=status_f)
    if not inc_upload:
        q = q.filter_by(upload_status="not_uploaded")
    if branch_id:
        q = q.filter_by(branch_id=branch_id)
    if month:
        try:
            parts = month.split("-")
            yr = int(parts[0])
            mo = int(parts[1]) if len(parts) > 1 else 0
            q = q.filter(func.extract("year", PaymentRequest.date) == yr)
            if mo:
                q = q.filter(func.extract("month", PaymentRequest.date) == mo)
        except (ValueError, IndexError):
            pass
    if not current_user.is_mds:
        q = q.filter_by(branch_id=current_user.branch_id)

    records = q.order_by(PaymentRequest.date.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Approved Payments"

    orange_fill  = PatternFill("solid", fgColor="F5821F")
    navy_fill    = PatternFill("solid", fgColor="0B1E3D")
    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    body_font    = Font(name="Calibri", size=10)
    thin = Side(style="thin", color="1E3F7A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    headers = [
        "Reference","Date","Branch","Category","Description",
        "Qty","Rate (₦)","Requested (₦)","Approved (₦)",
        "Beneficiary Name","Account Number","Bank","Bank Code",
        "MDS Comment","Upload Status",
    ]
    ws.append(["Sure Diagnostics — Payment Export"])
    ws.append([f"Generated: {datetime.utcnow().strftime('%d %b %Y %H:%M')} UTC"])
    ws.append([])
    ws.append(headers)

    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="F5821F")
    ws["A2"].font = Font(name="Calibri", size=9, color="7A9CC4")

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = orange_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    for pr in records:
        ws.append([
            pr.reference,
            pr.date.strftime("%d/%m/%Y"),
            pr.branch.name   if pr.branch   else "",
            pr.category.name if pr.category else "",
            pr.description,
            pr.quantity,
            float(pr.rate),
            float(pr.requested_amount),
            float(pr.approved_amount) if pr.approved_amount else "",
            pr.beneficiary_name,
            pr.beneficiary_account,
            pr.beneficiary_bank,
            pr.bank_code or "",
            pr.mds_comment or "",
            pr.upload_status.replace("_", " ").title(),
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"SureDiagnostics_Payments_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
