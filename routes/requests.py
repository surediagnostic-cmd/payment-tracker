from collections import defaultdict
from datetime import datetime, date
from decimal import Decimal
from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app
from flask_login import login_required, current_user
from sqlalchemy import func
from sqlalchemy.orm import subqueryload, joinedload
from app import db
from models import PaymentRequest, PaymentRequestItem, Branch, Category, User
from utils import send_email

requests_bp = Blueprint("requests", __name__)


def _parse_month(month_str):
    now = datetime.utcnow()
    if month_str:
        try:
            parts = month_str.split("-")
            yr = int(parts[0])
            mo = int(parts[1]) if len(parts) > 1 else 0
            return yr, mo
        except Exception:
            pass
    return now.year, now.month


def _eager_pr():
    """Return a base query that pre-loads all lazy relationships."""
    return PaymentRequest.query.options(
        subqueryload(PaymentRequest.items).joinedload(PaymentRequestItem.category),
        joinedload(PaymentRequest.branch),
        joinedload(PaymentRequest.submitter),
    )


@requests_bp.route("/dashboard")
@login_required
def dashboard():
    try:
        return _dashboard_inner()
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        print(f"[dashboard error]: {e}", flush=True)
        # Return plain HTML — avoids Jinja2/DB dependency in the fallback
        return (
            f"<html><body style='font-family:sans-serif;padding:40px;"
            f"background:#0b1e3d;color:#e8edf5;'>"
            f"<h2 style='color:#f5821f;'>Dashboard temporarily unavailable</h2>"
            f"<p style='color:#7a9cc4;margin:12px 0;'>Error: {str(e)}</p>"
            f"<a href='/dashboard' style='color:#ff9d45;margin-right:20px;'>↺ Retry</a>"
            f"<a href='/requests' style='color:#ff9d45;'>All Payments →</a>"
            f"</body></html>"
        ), 200


def _dashboard_inner():
    month_str      = request.args.get("month", "")
    selected_branch = request.args.get("branch_id", type=int)
    yr, mo = _parse_month(month_str)
    month_label = date(yr, mo, 1).strftime("%B %Y") if mo else f"All of {yr}"

    branches = Branch.query.filter_by(is_active=True).order_by(Branch.name).all()

    def month_filter(q):
        q = q.filter(func.extract("year", PaymentRequest.date) == yr)
        if mo:
            q = q.filter(func.extract("month", PaymentRequest.date) == mo)
        return q

    if current_user.is_mds:
        # ── Pending (no month filter — show all pending) ─────────────────────
        pending_q = _eager_pr().filter_by(status="pending").order_by(PaymentRequest.created_at.desc())
        if selected_branch:
            pending_q = pending_q.filter_by(branch_id=selected_branch)
        pending = pending_q.all()

        # ── Approved this period ──────────────────────────────────────────────
        approved_q = PaymentRequest.query.filter_by(status="approved")
        approved_q = month_filter(approved_q)
        if selected_branch:
            approved_q = approved_q.filter_by(branch_id=selected_branch)
        approved_this_month = approved_q.all()
        total_approved = sum(float(r.approved_amount or 0) for r in approved_this_month)
        total_pending_amt = sum(
            float(r.requested_amount)
            for r in PaymentRequest.query.filter_by(status="pending").all()
        )

        # ── Approval rate ─────────────────────────────────────────────────────
        reviewed = PaymentRequest.query.filter(PaymentRequest.status.in_(["approved", "rejected"]))
        reviewed = month_filter(reviewed)
        if selected_branch:
            reviewed = reviewed.filter_by(branch_id=selected_branch)
        reviewed_count = reviewed.count()
        approved_count = len(approved_this_month)
        approval_rate  = round((approved_count / reviewed_count) * 100) if reviewed_count else 0

        # ── Status counts ─────────────────────────────────────────────────────
        def status_count(s):
            q = PaymentRequest.query.filter_by(status=s)
            q = month_filter(q)
            if selected_branch:
                q = q.filter_by(branch_id=selected_branch)
            return q.count()

        uploaded_q = month_filter(PaymentRequest.query.filter_by(upload_status="uploaded"))
        if selected_branch:
            uploaded_q = uploaded_q.filter_by(branch_id=selected_branch)

        status_counts = {
            "pending":  status_count("pending"),
            "approved": status_count("approved"),
            "rejected": status_count("rejected"),
            "uploaded": uploaded_q.count(),
        }

        # ── Branch totals ─────────────────────────────────────────────────────
        branch_totals_q = db.session.query(
            Branch.name.label("branch"),
            func.sum(PaymentRequest.requested_amount).label("requested"),
            func.sum(PaymentRequest.approved_amount).label("approved"),
            func.count(PaymentRequest.id).label("count"),
        ).select_from(PaymentRequest)\
         .join(Branch, Branch.id == PaymentRequest.branch_id)\
         .filter(PaymentRequest.status.in_(["approved", "pending"]))
        branch_totals_q = month_filter(branch_totals_q)
        if selected_branch:
            branch_totals_q = branch_totals_q.filter(PaymentRequest.branch_id == selected_branch)
        branch_totals = branch_totals_q.group_by(Branch.name)\
            .order_by(func.sum(PaymentRequest.approved_amount).desc()).all()

        # ── Per-branch request details (for accordion) ────────────────────────
        all_br_q = _eager_pr().filter(PaymentRequest.status.in_(["approved", "pending"]))
        all_br_q = month_filter(all_br_q)
        if selected_branch:
            all_br_q = all_br_q.filter_by(branch_id=selected_branch)
        branch_requests = defaultdict(list)
        for r in all_br_q.order_by(PaymentRequest.branch_id, PaymentRequest.date.desc()).all():
            branch_requests[r.branch.name].append(r)

        # ── Category breakdown via line items ─────────────────────────────────
        cat_q = db.session.query(
            Category.name,
            Category.cost_type,
            func.sum(PaymentRequestItem.amount).label("total"),
            func.count(PaymentRequestItem.id).label("count"),
        ).select_from(PaymentRequestItem)\
         .join(Category, Category.id == PaymentRequestItem.category_id)\
         .join(PaymentRequest, PaymentRequest.id == PaymentRequestItem.request_id)\
         .filter(PaymentRequest.status == "approved")
        cat_q = month_filter(cat_q)
        if selected_branch:
            cat_q = cat_q.filter(PaymentRequest.branch_id == selected_branch)
        category_data = cat_q.group_by(Category.name, Category.cost_type)\
                             .order_by(func.sum(PaymentRequestItem.amount).desc()).all()

        # ── Direct cost / overhead totals ─────────────────────────────────────
        def _cost_total(cost_type):
            q = db.session.query(func.sum(PaymentRequestItem.amount))\
                .select_from(PaymentRequestItem)\
                .join(Category, Category.id == PaymentRequestItem.category_id)\
                .join(PaymentRequest, PaymentRequest.id == PaymentRequestItem.request_id)\
                .filter(PaymentRequest.status == "approved", Category.cost_type == cost_type)
            q = month_filter(q)
            if selected_branch:
                q = q.filter(PaymentRequest.branch_id == selected_branch)
            return float(q.scalar() or 0)

        total_direct_cost = _cost_total("direct_cost")
        total_overhead    = _cost_total("overhead")

        # ── Variance ──────────────────────────────────────────────────────────
        var_q = _eager_pr().filter(
            PaymentRequest.status == "approved",
            PaymentRequest.approved_amount != PaymentRequest.requested_amount,
        )
        var_q = month_filter(var_q)
        if selected_branch:
            var_q = var_q.filter_by(branch_id=selected_branch)
        variance_data = var_q.order_by(PaymentRequest.date.desc()).limit(30).all()

        # ── Recent ────────────────────────────────────────────────────────────
        recent_q = _eager_pr()
        recent_q = month_filter(recent_q)
        if selected_branch:
            recent_q = recent_q.filter_by(branch_id=selected_branch)
        recent = recent_q.order_by(PaymentRequest.created_at.desc()).limit(20).all()

        effective_month_str = month_str or (f"{yr:04d}-{mo:02d}" if mo else f"{yr:04d}-0")
        return render_template(
            "dashboard.html",
            month_str=effective_month_str, month_label=month_label,
            selected_branch=selected_branch, branches=branches,
            pending=pending, total_approved=total_approved,
            total_pending_amt=total_pending_amt, approval_rate=approval_rate,
            status_counts=status_counts, branch_totals=branch_totals,
            branch_requests=branch_requests,
            total_direct_cost=total_direct_cost, total_overhead=total_overhead,
            category_data=category_data, variance_data=variance_data, recent=recent,
        )

    else:
        # ── Accountant dashboard ──────────────────────────────────────────────
        user_branches    = current_user.branches
        user_branch_ids  = [b.id for b in user_branches]

        # Branch filter (relevant when accountant assigned to multiple branches)
        if selected_branch and selected_branch in user_branch_ids:
            acct_branch_filter = selected_branch
        else:
            acct_branch_filter = None

        my_q = _eager_pr().filter_by(submitted_by=current_user.id)
        my_q = month_filter(my_q)
        if acct_branch_filter:
            my_q = my_q.filter_by(branch_id=acct_branch_filter)
        my_requests = my_q.order_by(PaymentRequest.created_at.desc()).all()

        my_pending  = sum(1 for r in my_requests if r.status == "pending")
        my_approved = [r for r in my_requests if r.status == "approved"]
        my_total    = sum(float(r.approved_amount or 0) for r in my_approved)

        cat_q = db.session.query(
            Category.name,
            Category.cost_type,
            func.sum(PaymentRequestItem.amount).label("total"),
            func.count(PaymentRequestItem.id).label("count"),
        ).select_from(PaymentRequestItem)\
         .join(Category, Category.id == PaymentRequestItem.category_id)\
         .join(PaymentRequest, PaymentRequest.id == PaymentRequestItem.request_id)\
         .filter(PaymentRequest.status == "approved",
                 PaymentRequest.submitted_by == current_user.id)
        cat_q = month_filter(cat_q)
        if acct_branch_filter:
            cat_q = cat_q.filter(PaymentRequest.branch_id == acct_branch_filter)
        category_data = cat_q.group_by(Category.name, Category.cost_type)\
                             .order_by(func.sum(PaymentRequestItem.amount).desc()).all()

        status_counts = {
            "pending":  sum(1 for r in my_requests if r.status == "pending"),
            "approved": sum(1 for r in my_requests if r.status == "approved"),
            "rejected": sum(1 for r in my_requests if r.status == "rejected"),
            "uploaded": sum(1 for r in my_requests if r.upload_status == "uploaded"),
        }

        effective_month_str = month_str or (f"{yr:04d}-{mo:02d}" if mo else f"{yr:04d}-0")
        return render_template(
            "dashboard.html",
            month_str=effective_month_str, month_label=month_label,
            selected_branch=acct_branch_filter, branches=branches,
            user_branches=user_branches,
            my_requests=my_requests, my_pending=my_pending,
            my_total=my_total, approved_count=len(my_approved),
            category_data=category_data, status_counts=status_counts,
        )


@requests_bp.route("/requests/new", methods=["GET", "POST"])
@login_required
def new_request():
    if current_user.is_mds:
        flash("MDS account cannot submit payment requests.", "error")
        return redirect(url_for("requests.dashboard"))

    categories = Category.query.filter_by(is_active=True).order_by(Category.name).all()
    user_branch_list = current_user.branches

    if not user_branch_list:
        flash("Your account has no branch assigned. Contact MDS Admin.", "error")
        return redirect(url_for("requests.dashboard"))

    if request.method == "POST":
        try:
            branch_id = int(request.form["branch_id"])
            valid_branch_ids = [b.id for b in current_user.branches]
            if branch_id not in valid_branch_ids:
                flash("Invalid branch selection.", "error")
                return redirect(url_for("requests.new_request"))

            descriptions = request.form.getlist("descriptions[]")
            category_ids = request.form.getlist("category_ids[]")
            quantities   = request.form.getlist("quantities[]")
            rates        = request.form.getlist("rates[]")

            if not descriptions or not any(d.strip() for d in descriptions):
                flash("At least one line item is required.", "error")
                return redirect(url_for("requests.new_request"))

            items_data = []
            total = Decimal(0)
            for i in range(len(descriptions)):
                desc = descriptions[i].strip()
                if not desc:
                    continue
                qty    = int(quantities[i])
                rate   = Decimal(rates[i].replace(",", ""))
                amount = qty * rate
                total += amount
                items_data.append({
                    "description": desc,
                    "category_id": int(category_ids[i]),
                    "quantity":    qty,
                    "rate":        rate,
                    "amount":      amount,
                })

            ref = PaymentRequest.generate_reference(branch_id)
            pr = PaymentRequest(
                reference=ref,
                date=datetime.strptime(request.form["date"], "%Y-%m-%d").date(),
                branch_id=branch_id,
                beneficiary_name=request.form["beneficiary_name"].strip(),
                beneficiary_account=request.form["beneficiary_account"].strip(),
                beneficiary_bank=request.form["beneficiary_bank"].strip(),
                bank_code=request.form.get("bank_code", "").strip(),
                requested_amount=total,
                submitted_by=current_user.id,
            )
            db.session.add(pr)
            db.session.flush()

            for item in items_data:
                db.session.add(PaymentRequestItem(request_id=pr.id, **item))

            db.session.commit()

            mds_email = current_app.config.get("MDS_EMAIL")
            if mds_email:
                item_lines = "\n".join(
                    f"  • {it['description']} — ₦{it['amount']:,.2f}" for it in items_data
                )
                send_email(
                    to=mds_email,
                    subject=f"[Sure Finance] New Payment Request — {pr.reference}",
                    body=(
                        f"A new payment request has been submitted.\n\n"
                        f"Reference: {pr.reference}\n"
                        f"Branch: {pr.branch.name}\n"
                        f"Total Requested: ₦{total:,.2f}\n"
                        f"Beneficiary: {pr.beneficiary_name} ({pr.beneficiary_bank})\n"
                        f"Submitted by: {current_user.name}\n\n"
                        f"Items:\n{item_lines}\n\n"
                        f"Please log in to review this request."
                    ),
                )

            flash(f"Payment request {ref} submitted successfully.", "success")
            return redirect(url_for("requests.dashboard"))
        except Exception as e:
            try:
                db.session.rollback()
            except Exception:
                pass
            print(f"[submit error]: {e}", flush=True)
            flash(f"Error submitting request: {str(e)}", "error")

    return render_template("new_request.html", categories=categories, user_branches=user_branch_list)


@requests_bp.route("/requests/<int:req_id>")
@login_required
def view_request(req_id):
    pr = _eager_pr().filter_by(id=req_id).first_or_404()
    if not current_user.is_mds and pr.submitted_by != current_user.id:
        flash("Access denied.", "error")
        return redirect(url_for("requests.dashboard"))
    return render_template("request_detail.html", pr=pr)


@requests_bp.route("/requests/<int:req_id>/upload", methods=["POST"])
@login_required
def mark_uploaded(req_id):
    pr = PaymentRequest.query.get_or_404(req_id)
    if pr.status != "approved":
        flash("Only approved requests can be marked as uploaded.", "error")
        return redirect(url_for("requests.view_request", req_id=req_id))

    pr.upload_status = "uploaded"
    db.session.commit()
    flash(f"{pr.reference} marked as uploaded.", "success")
    return redirect(url_for("requests.view_request", req_id=req_id))


@requests_bp.route("/requests/template")
@login_required
def download_template():
    """Download a blank Excel payment request template (one row = one request)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file

    # Fetch live category/branch names for the Instructions sheet
    cats     = [c.name for c in Category.query.filter_by(is_active=True).order_by(Category.name).all()]
    branches = [b.name for b in Branch.query.filter_by(is_active=True).order_by(Branch.name).all()]

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Payment Requests"
    navy, lt, grey = "0B1E3D", "E8EDF5", "F2F4F7"

    headers    = ["Date (YYYY-MM-DD)", "Branch", "Beneficiary Name",
                  "Account Number", "Bank", "Bank Code (opt)",
                  "Description", "Category", "Quantity", "Rate (₦)", "Amount (₦)"]
    col_widths = [18, 14, 24, 16, 18, 14, 32, 24, 10, 13, 13]

    thin = Side(style="thin", color="C5CDD8")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font      = Font(bold=True, color=lt, name="Calibri", size=10)
        cell.fill      = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = bdr
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 28

    # Sample row (row 2) — shown in italic grey so users know to replace it
    sample = ["2026-05-01", branches[0] if branches else "Ijofi",
              "Vendor Name", "0123456789", "Zenith Bank", "057",
              "Example: Lab reagents batch", cats[0] if cats else "Lab Supplies / Reagents",
              1, 10000, "=J2*I2"]
    for i, v in enumerate(sample, 1):
        cell = ws.cell(row=2, column=i, value=v)
        cell.fill      = PatternFill("solid", fgColor=grey)
        cell.font      = Font(name="Calibri", size=10, italic=True, color="888888")
        cell.alignment = Alignment(horizontal="center")
        cell.border    = bdr

    # Empty data rows 3–51
    for r in range(3, 52):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c, value=None)
            cell.border    = bdr
            cell.alignment = Alignment(horizontal="center")
            cell.font      = Font(name="Calibri", size=10)
        # Auto-formula for Amount column (col 11)
        ws.cell(row=r, column=11, value=f"=J{r}*I{r}")

    ws.freeze_panes = "A2"

    # ── Instructions sheet ───────────────────────────────────────────────
    wi = wb.create_sheet("Instructions")
    lines = [
        ("Sure Diagnostics — Bulk Payment Request Template", True, 13),
        ("", False, 10),
        ("HOW TO USE", True, 11),
        ("1. Each row is ONE complete payment request (with one line item).", False, 10),
        ("2. Delete the grey sample row (row 2) before uploading.", False, 10),
        ("3. Date format must be YYYY-MM-DD   e.g. 2026-05-04", False, 10),
        ("4. Amount auto-calculates (Quantity × Rate). You can leave it.", False, 10),
        ("5. Save as .xlsx then upload via 'Bulk Upload from Template'.", False, 10),
        ("", False, 10),
        ("VALID BRANCHES", True, 11),
        ("  " + "  |  ".join(branches), False, 10),
        ("", False, 10),
        ("VALID CATEGORIES (copy exactly)", True, 11),
    ]
    for cat in cats:
        lines.append(("  " + cat, False, 10))

    for row, (text, bold, size) in enumerate(lines, 1):
        cell = wi.cell(row=row, column=1, value=text)
        cell.font = Font(bold=bold, size=size, name="Calibri",
                         color="1A3D7C" if bold else "333333")
    wi.column_dimensions["A"].width = 70

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True,
                     download_name="Sure_Diagnostics_Payment_Template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@requests_bp.route("/requests/bulk-upload", methods=["GET", "POST"])
@login_required
def bulk_upload():
    """Import multiple payment requests from the filled-in Excel template."""
    if current_user.is_mds:
        flash("MDS account cannot submit payment requests.", "error")
        return redirect(url_for("requests.dashboard"))

    user_branches    = current_user.branches
    user_branch_ids  = {b.id for b in user_branches}

    if request.method == "POST":
        file = request.files.get("excel_file")
        if not file or not file.filename.lower().endswith((".xlsx", ".xls")):
            flash("Please upload a valid Excel file (.xlsx).", "error")
            return redirect(url_for("requests.bulk_upload"))

        try:
            from openpyxl import load_workbook
            wb = load_workbook(file, data_only=True)
            ws = wb.active

            all_cats     = {c.name.lower().strip(): c
                            for c in Category.query.filter_by(is_active=True).all()}
            all_branches = {b.name.lower().strip(): b
                            for b in Branch.query.filter_by(is_active=True).all()}

            created, skipped, errors = 0, 0, []

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                # Skip fully empty rows
                if not any(cell for cell in row if cell is not None):
                    skipped += 1
                    continue
                # Skip if fewer columns than expected
                if len(row) < 10:
                    errors.append(f"Row {row_num}: too few columns, skipped.")
                    continue

                date_val, branch_raw, bene_name, bene_acct, bene_bank, \
                    bank_code, description, cat_raw, qty_raw, rate_raw = \
                    (row[i] for i in range(10))

                # Skip the sample/example row
                if str(description or "").lower().startswith("example:"):
                    skipped += 1
                    continue

                # ── Validate required fields ──────────────────────────────
                missing = [f for f, v in [
                    ("Date", date_val), ("Branch", branch_raw),
                    ("Beneficiary Name", bene_name), ("Account Number", bene_acct),
                    ("Bank", bene_bank), ("Description", description),
                    ("Category", cat_raw), ("Quantity", qty_raw), ("Rate", rate_raw),
                ] if not v]
                if missing:
                    errors.append(f"Row {row_num}: missing {', '.join(missing)}.")
                    continue

                # ── Parse date ────────────────────────────────────────────
                try:
                    from datetime import date as dt_date
                    if isinstance(date_val, dt_date):
                        req_date = date_val
                    else:
                        req_date = datetime.strptime(str(date_val).strip(), "%Y-%m-%d").date()
                except ValueError:
                    errors.append(f"Row {row_num}: invalid date '{date_val}' — use YYYY-MM-DD.")
                    continue

                # ── Lookup branch ─────────────────────────────────────────
                branch = all_branches.get(str(branch_raw).lower().strip())
                if not branch:
                    errors.append(f"Row {row_num}: unknown branch '{branch_raw}'.")
                    continue
                if branch.id not in user_branch_ids:
                    errors.append(f"Row {row_num}: branch '{branch.name}' not assigned to you.")
                    continue

                # ── Lookup category ───────────────────────────────────────
                cat_key = str(cat_raw).lower().strip()
                cat = all_cats.get(cat_key)
                if not cat:  # fuzzy: first category whose name contains the key
                    cat = next((c for c in all_cats.values()
                                if cat_key in c.name.lower()), None)
                if not cat:
                    errors.append(f"Row {row_num}: unknown category '{cat_raw}'.")
                    continue

                # ── Parse numbers ─────────────────────────────────────────
                try:
                    qty  = int(float(qty_raw))
                    rate = Decimal(str(float(rate_raw)))
                    if qty <= 0 or rate <= 0:
                        raise ValueError("must be positive")
                    amount = qty * rate
                except (ValueError, TypeError) as num_err:
                    errors.append(f"Row {row_num}: invalid qty/rate — {num_err}.")
                    continue

                # ── Create request + item ─────────────────────────────────
                try:
                    ref = PaymentRequest.generate_reference(branch.id)
                    pr  = PaymentRequest(
                        reference         = ref,
                        date              = req_date,
                        branch_id         = branch.id,
                        beneficiary_name  = str(bene_name).strip(),
                        beneficiary_account = str(bene_acct).strip(),
                        beneficiary_bank  = str(bene_bank).strip(),
                        bank_code         = str(bank_code).strip() if bank_code else "",
                        requested_amount  = amount,
                        submitted_by      = current_user.id,
                    )
                    db.session.add(pr)
                    db.session.flush()
                    db.session.add(PaymentRequestItem(
                        request_id  = pr.id,
                        description = str(description).strip(),
                        category_id = cat.id,
                        quantity    = qty,
                        rate        = rate,
                        amount      = amount,
                    ))
                    created += 1
                except Exception as row_err:
                    db.session.rollback()
                    errors.append(f"Row {row_num}: {row_err}.")
                    continue

            if created:
                db.session.commit()
                flash(f"{created} payment request(s) submitted successfully.", "success")
            else:
                db.session.rollback()
                flash("No valid rows found — nothing was created.", "warning")

            for err in errors[:8]:
                flash(err, "warning")

            return redirect(url_for("requests.dashboard"))

        except Exception as e:
            try:
                db.session.rollback()
            except Exception:
                pass
            flash(f"Could not read the file: {str(e)}", "error")
            return redirect(url_for("requests.bulk_upload"))

    return render_template("bulk_upload.html", user_branches=user_branches)


@requests_bp.route("/requests/<int:req_id>/delete", methods=["POST"])
@login_required
def delete_request(req_id):
    if not current_user.is_mds:
        flash("Only MDS can delete payment requests.", "error")
        return redirect(url_for("requests.dashboard"))
    pr = PaymentRequest.query.get_or_404(req_id)
    ref = pr.reference
    db.session.delete(pr)
    db.session.commit()
    flash(f"Payment request {ref} has been permanently deleted.", "warning")
    return redirect(url_for("requests.list_requests"))


@requests_bp.route("/requests")
@login_required
def list_requests():
    q = _eager_pr()
    if not current_user.is_mds:
        q = q.filter_by(submitted_by=current_user.id)

    branch_id = request.args.get("branch_id", type=int)
    month     = request.args.get("month", "")
    status    = request.args.get("status", "")

    if branch_id: q = q.filter_by(branch_id=branch_id)
    if status:    q = q.filter_by(status=status)
    if month:
        try:
            parts = month.split("-")
            yr = int(parts[0]); mo = int(parts[1]) if len(parts) > 1 else 0
            q = q.filter(func.extract("year", PaymentRequest.date) == yr)
            if mo:
                q = q.filter(func.extract("month", PaymentRequest.date) == mo)
        except ValueError:
            pass

    requests_list = q.order_by(PaymentRequest.created_at.desc()).all()
    branches = Branch.query.filter_by(is_active=True).all()

    return render_template(
        "list_requests.html",
        requests=requests_list, branches=branches,
        filters={"branch_id": branch_id, "month": month, "status": status},
    )
